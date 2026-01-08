"""
Google Maps Scraper Module
Refactored for web API usage
"""
import asyncio
import inspect
import os
import re
from datetime import datetime
from typing import Callable, Optional, Dict, List, Set, Tuple

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


class GoogleMapsScraper:
    def __init__(self, progress_callback: Optional[Callable] = None):
        """
        Initialize scraper with optional progress callback

        Args:
            progress_callback: Function to call with progress updates
                             Signature: callback(event_type, data)
        """
        self.progress_callback = progress_callback
        self.is_running = False
        self.is_paused = False
        self.results: List[Dict] = []
        self.base_folder = "resultados"
        self.retry_count = 3
        self._seen_keys: Set[Tuple[str, str]] = set()

        # Departamentos de Perú
        self.departamentos_peru = [
            "Lima", "Arequipa", "Cusco", "Trujillo", "Chiclayo", "Piura", "Iquitos",
            "Huancayo", "Tacna", "Ica", "Juliaca", "Pucallpa", "Cajamarca", "Ayacucho",
            "Huánuco", "Chimbote", "Tarapoto", "Tumbes", "Puno", "Sullana", "Chincha Alta",
            "Huaraz", "Talara", "Jaén", "Abancay"
        ]

        # Sinónimos para búsqueda inteligente
        self.sinonimos = {
            "restaurante": ["restaurant", "comida", "food", "gastronomía", "cocina"],
            "hotel": ["hospedaje", "alojamiento", "hostal", "lodge", "inn"],
            "minería": ["mining", "minera", "mina", "extracción minera"],
            "construcción": ["construccion", "building", "obra", "contractor"],
            "farmacia": ["pharmacy", "botica", "drugstore"],
            "panadería": ["bakery", "pan", "bread", "pastelería"],
            "ferretería": ["hardware store", "herramientas", "tools"],
        }

        if not os.path.exists(self.base_folder):
            os.makedirs(self.base_folder)

    def _emit_progress(self, event_type: str, data: Dict):
        """Emit progress event to callback"""
        if self.progress_callback:
            result = self.progress_callback(event_type, data)
            if inspect.isawaitable(result):
                asyncio.create_task(result)

    def _normalize_text(self, text: str) -> str:
        """Normalize text for comparison"""
        replacements = {
            'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'ñ': 'n'
        }
        text = text.lower()
        for old, new in replacements.items():
            text = text.replace(old, new)
        return text

    def _make_key(self, nombre: str, direccion: str) -> Tuple[str, str]:
        """Create a normalized key to avoid duplicates"""
        def norm(s: str) -> str:
            s = s.lower()
            s = re.sub(r"\s+", " ", s)
            return s.strip()
        return (norm(nombre), norm(direccion))

    def expand_search_term(self, rubro: str) -> List[str]:
        """Expand search with synonyms"""
        rubro_norm = self._normalize_text(rubro)

        for key, synonyms in self.sinonimos.items():
            key_norm = self._normalize_text(key)
            synonyms_norm = [self._normalize_text(s) for s in synonyms]
            if rubro_norm == key_norm or rubro_norm in synonyms_norm or rubro_norm in key_norm:
                ordered = [rubro.strip()] + synonyms
                unique = []
                seen = set()
                for term in ordered:
                    normalized = self._normalize_text(term)
                    if normalized not in seen:
                        seen.add(normalized)
                        unique.append(term)
                return unique[:3]

        return [rubro]

    def _build_location_terms(self, departamento: str, pais: str) -> List[str]:
        """Build location terms; for Lima, expand to districts."""
        departamento_norm = departamento.strip().lower()
        if departamento_norm == "lima":
            distritos_lima = [
                "Lima", "Miraflores", "San Isidro", "Santiago de Surco", "La Molina", "San Borja",
                "Barranco", "San Miguel", "Pueblo Libre", "Jesus Maria", "Magdalena del Mar",
                "Lince", "Breña", "Rimac", "San Luis", "Chorrillos", "Ate", "Callao",
                "Los Olivos", "San Martin de Porres", "Independencia", "Comas",
                "Villa El Salvador", "Villa Maria del Triunfo", "San Juan de Lurigancho",
                "San Juan de Miraflores", "El Agustino", "Santa Anita", "La Victoria",
                "Carabayllo", "Puente Piedra", "Surquillo"
            ]
            return [f"{distrito}, Lima, {pais}" for distrito in distritos_lima]
        return [f"{departamento}, {pais}"]

    async def scrape(self, rubro: str, departamento: str, pais: str, cantidad: Optional[int],
                    headless: bool = True, expanded_search: bool = True):
        """Main scraping method"""
        self.is_running = True
        self.is_paused = False
        self.results = []
        self._seen_keys.clear()

        search_terms = self.expand_search_term(rubro) if expanded_search else [rubro]
        location_terms = self._build_location_terms(departamento, pais)
        # Si no se requiere límite estricto, usamos un número alto para recorrer todo
        effective_limit = cantidad if cantidad and cantidad > 0 else 1_000_000

        self._emit_progress("log", {
            "message": f"Iniciando busqueda: {rubro} en {departamento}, {pais}",
            "type": "info"
        })

        if len(search_terms) > 1:
            self._emit_progress("log", {
                "message": f"Busqueda inteligente: {', '.join(search_terms[:3])}",
                "type": "info"
            })

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=headless,
                args=["--disable-blink-features=AutomationControlled"],
            )

            context = await browser.new_context(
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            )
            context.set_default_timeout(60000)
            page = await context.new_page()

            for loc_idx, location in enumerate(location_terms, 1):
                if not self.is_running:
                    break

                for term_idx, search_term in enumerate(search_terms, 1):
                    if not self.is_running:
                        break

                    await self._wait_if_paused()

                    if len(search_terms) > 1 or len(location_terms) > 1:
                        self._emit_progress("log", {
                            "message": f"Termino {term_idx}/{len(search_terms)} en zona {loc_idx}/{len(location_terms)}: {search_term} @ {location}",
                            "type": "info"
                        })

                    search_query = f"{search_term} en {location}"
                    url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"

                    for attempt in range(1, self.retry_count + 1):
                        try:
                            self._emit_progress("log", {
                                "message": f"Cargando pagina (intento {attempt})...",
                                "type": "info"
                            })
                            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
                            await asyncio.sleep(3)
                            self._emit_progress("log", {
                                "message": "Pagina cargada exitosamente",
                                "type": "success"
                            })
                            break
                        except Exception as e:
                            if attempt < self.retry_count:
                                await asyncio.sleep(2)
                            else:
                                self._emit_progress("log", {
                                    "message": f"Error cargando pagina: {str(e)}",
                                    "type": "error"
                                })

                    await self._scroll_extensive(page, effective_limit)

                    if not self.is_running:
                        break

                    await self._extract_business_data(page, effective_limit, departamento)

            await browser.close()

        self._emit_progress("log", {
            "message": f"Scraping completado: {len(self.results)} negocios extraidos",
            "type": "success"
        })

        self.is_running = False
        return self.results

    async def _scroll_extensive(self, page, max_results: int):
        """Scroll to load more results"""
        self._emit_progress("log", {
            "message": "Cargando resultados...",
            "type": "info"
        })

        try:
            await page.wait_for_selector('div[role="feed"]', timeout=30000)

            previous = 0
            no_change_count = 0
            max_no_change = 8
            scroll_count = 0
            max_scrolls = 100

            while scroll_count < max_scrolls and no_change_count < max_no_change and self.is_running:
                current = await page.locator('div[role="feed"] > div').count()

                await self._wait_if_paused()

                if current >= max_results:
                    self._emit_progress("log", {
                        "message": f"Suficientes resultados cargados: {current}",
                        "type": "success"
                    })
                    break

                await page.evaluate(
                    """
                    const feed = document.querySelector('div[role="feed"]');
                    if (feed) {
                        feed.scrollTop = feed.scrollHeight;
                        setTimeout(() => {
                            feed.scrollTop = feed.scrollHeight + 1000;
                        }, 100);
                    }
                    """
                )

                await asyncio.sleep(1.2)
                scroll_count += 1

                if current == previous:
                    no_change_count += 1
                else:
                    no_change_count = 0
                    previous = current
                    if current % 10 == 0:
                        self._emit_progress("log", {
                            "message": f"Cargados: {current} resultados",
                            "type": "info"
                        })

            final_count = await page.locator('div[role="feed"] > div').count()
            self._emit_progress("log", {
                "message": f"Total encontrados: {final_count} resultados",
                "type": "success"
            })

        except Exception as e:
            self._emit_progress("log", {
                "message": f"Error en scroll: {str(e)}",
                "type": "error"
            })

    async def _extract_business_data(self, page, max_results: int, departamento: str):
        """Extract business data from loaded results"""
        self._emit_progress("log", {
            "message": "Extrayendo datos de negocios...",
            "type": "info"
        })

        links = await page.locator('a[href*="/maps/place/"]').all()

        remaining = max_results - len(self.results)
        if remaining <= 0:
            return

        total = min(len(links), remaining)
        self._emit_progress("log", {
            "message": f"Procesando {total} negocios con extraccion paralela...",
            "type": "info"
        })

        # Slicing the links we actually need
        target_links = links[:total]
        
        # Concurrency control (e.g. 5 tabs at once)
        sem = asyncio.Semaphore(5)

        async def process_link(link_element):
            if not self.is_running:
                return

            async with sem:
                # Need to be robust with element handles in parallel
                try:
                    # We open a NEW page for each result to avoid navigating back and forth on the main page
                    # However, clicking opens in the same tab usually on Maps OR opens a side panel.
                    # Best approach for Maps seeking speed: 
                    # Get the HREF, open in new tab.
                    
                    href = await link_element.get_attribute("href")
                    if not href:
                        return

                    new_page = await page.context.new_page()
                    
                    try:
                        await new_page.goto(href, wait_until="domcontentloaded", timeout=45000)
                        
                        # Random sleep to be polite but fast
                        await asyncio.sleep(1.5)
                        
                        data = await self._get_business_details(new_page, departamento)
                        if data and data.get("nombre"):
                            key = self._make_key(data["nombre"], data.get("direccion", ""))
                            if key not in self._seen_keys:
                                self._seen_keys.add(key)
                                self.results.append(data)

                                self._emit_progress("result", {
                                    "data": data,
                                    "index": len(self.results)
                                })
                                
                                # Update progress percentage
                                current_total = len(self.results)
                                self._emit_progress("progress", {
                                    "current": current_total,
                                    "total": max_results,
                                    "percentage": int((current_total / max_results) * 100)
                                })

                                if len(self.results) % 10 == 0:
                                    self._emit_progress("log", {
                                        "message": f"Avance: {len(self.results)} extraidos",
                                        "type": "info"
                                    })
                    finally:
                        await new_page.close()
                except Exception as e:
                    # Silent fail for individual errors to keep momentum
                    pass

        # Launch all tasks
        tasks = [process_link(link) for link in target_links]
        await asyncio.gather(*tasks)

        self._emit_progress("log", {
            "message": f"Lote completado. Total actual: {len(self.results)}",
            "type": "success"
        })

    async def _get_business_details(self, page, departamento: str = "") -> Optional[Dict]:
        """Extract details from a single business"""
        data = {
            "nombre": "",
            "direccion": "",
            "telefono": "",
            "rating": "",
            "reviews": "",
            "estado": "",
            "correo": "",
            "sitio_web": "",
            "departamento": departamento,
        }

        try:
            try:
                data["nombre"] = await page.locator("h1.DUwDvf").first.inner_text(timeout=8000)
            except:
                pass

            try:
                addr = await page.locator('button[data-item-id="address"]').first.get_attribute("aria-label", timeout=8000)
                if addr:
                    data["direccion"] = addr.replace("Dirección: ", "").replace("Direcci\u00f3n: ", "").replace("Address: ", "")
            except:
                pass

            try:
                phone = await page.locator('button[data-item-id^="phone:tel:"]').first.get_attribute("aria-label", timeout=8000)
                if phone:
                    phone = phone.replace("Teléfono: ", "").replace("Phone: ", "")
                    data["telefono"] = phone
            except:
                pass

            try:
                rating = await page.locator('span[aria-label*="estrellas"]').first.text_content(timeout=5000)
                if rating:
                    data["rating"] = rating.strip()
            except:
                pass

            try:
                reviews = await page.locator('span[aria-label*="reseñas"]').first.text_content(timeout=5000)
                if reviews:
                    reviews = re.sub(r"\D", "", reviews)
                    data["reviews"] = reviews
            except:
                pass

            try:
                estado = await page.locator('div[role="button"][data-item-id="oh"] div[aria-label]').first.get_attribute("aria-label", timeout=5000)
                if estado:
                    data["estado"] = estado
            except:
                pass

            try:
                website = await page.locator('a[data-item-id="authority"]').first.get_attribute("href", timeout=5000)
                if website:
                    data["sitio_web"] = website
            except:
                pass

            try:
                section_text = await page.locator('div[data-item-id="summary"]').all_text_contents()
                section_text = " ".join(section_text)
                emails = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", section_text)
                if emails:
                    valid = [e for e in emails if "google" not in e.lower()]
                    if valid:
                        data["correo"] = valid[0]
            except:
                pass

            return data if data["nombre"] else None
        except:
            return None

    async def _wait_if_paused(self):
        """Wait while paused"""
        while self.is_paused and self.is_running:
            await asyncio.sleep(0.4)

    def pause(self):
        """Pause scraping"""
        self.is_paused = True
        self._emit_progress("log", {
            "message": "Scraping pausado",
            "type": "info"
        })

    def resume(self):
        """Resume scraping"""
        self.is_paused = False
        self._emit_progress("log", {
            "message": "Scraping reanudado",
            "type": "success"
        })

    def stop(self):
        """Stop scraping"""
        self.is_running = False
        self.is_paused = False
        self._emit_progress("log", {
            "message": "Scraping detenido",
            "type": "info"
        })

    def export_to_excel(self, rubro: str, departamento: str) -> str:
        """Export results to Excel file"""
        if not self.results:
            raise ValueError("No hay resultados para exportar")

        rubro_clean = rubro.strip().replace(" ", "_").replace("/", "-")
        rubro_folder = os.path.join(self.base_folder, rubro_clean)

        if not os.path.exists(rubro_folder):
            os.makedirs(rubro_folder)

        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
        dept_clean = departamento.strip().replace(" ", "_")
        filename = f"{rubro_clean}-{dept_clean}-{fecha}.xlsx"
        filepath = os.path.join(rubro_folder, filename)

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Resultados")
        else:
            ws.title = "Resultados"

        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )

        headers = ["#", "Nombre", "Direccion", "Telefono", "Rating", "Reviews", "Estado", "Correo", "Sitio Web"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for idx, row_data in enumerate(self.results, 2):
            ws.cell(row=idx, column=1, value=idx - 1).border = border
            ws.cell(row=idx, column=2, value=row_data["nombre"]).border = border
            ws.cell(row=idx, column=3, value=row_data["direccion"]).border = border
            ws.cell(row=idx, column=4, value=row_data["telefono"]).border = border
            ws.cell(row=idx, column=5, value=row_data.get("rating", "")).border = border
            ws.cell(row=idx, column=6, value=row_data.get("reviews", "")).border = border
            ws.cell(row=idx, column=7, value=row_data.get("estado", "")).border = border
            ws.cell(row=idx, column=8, value=row_data["correo"]).border = border
            ws.cell(row=idx, column=9, value=row_data["sitio_web"]).border = border

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 30
        ws.column_dimensions["I"].width = 40

        wb.save(filepath)

        self._emit_progress("log", {
            "message": f"Excel exportado: {filename}",
            "type": "success"
        })

        return filepath
